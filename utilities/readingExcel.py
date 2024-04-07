import openpyxl

workbook = openpyxl.load_workbook("..//excel//data//testdata.xlsx")
sheet = workbook["LoginTest"]
totalRows = sheet.max_row
totalCols = sheet.max_column

print("total rows are :{} and total columns are :{}".format(str(totalRows), str(totalCols)))

print(sheet.cell(row=2, column=1).value)

for rows in range(1, totalRows + 1):
    for cols in range(1, totalCols + 1):
        print(sheet.cell(row=rows, column=cols).value, end="        ")
    print()
