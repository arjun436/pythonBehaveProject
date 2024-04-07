import openpyxl

workbook = openpyxl.load_workbook("..//excel//data//testdata.xlsx")
sheet = workbook["LoginTest"]

#sheet.cell(row=1, column=3).value = "Age"
#
# workbook.save("..//excel//data//testdata.xlsx")

for rows in range(4, 8):
    for cols in range(1, 4):
        sheet.cell(row=rows, column=cols).value = "Test"

workbook.save("..//excel//data//testdata.xlsx")
