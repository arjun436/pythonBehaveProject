import allure
from allure_commons.types import AttachmentType
from selenium.webdriver.common.by import By
import pytest
import openpyxl

@pytest.fixture()
def log_on_failure(request, get_browser):
    driver = get_browser
    yield
    item = request.node
    if item.rep_call.failed:
        allure.attach(driver.get_screenshot_as_png(), name='failed', attachment_type=AttachmentType.PNG)


def get_data():
    return [
        ("arjun.miryala@gmail.com", "asdasdasds"),
        ("chirru.miryala@gmail.com", "gdgd"),
        ("andwit.miryala@gmail.com", "asdetr"),
    ]

def get_data2():
    workbook = openpyxl.load_workbook("..//excel//data//testdata.xlsx")
    sheet = workbook["LoginTest"]
    totalRows = sheet.max_row
    totalCols = sheet.max_column
    mainlist = []

    for i in range(2, totalRows+1):
        datalist=[]
        for j in range(1, totalCols+1):
            data = sheet.cell(row=i, column=j).value
            datalist.insert(j, data)
        mainlist.insert(i,datalist)
        print(mainlist)
    return mainlist


@pytest.fixture()
def teardown_function(get_browser):
    driver = get_browser
    driver.close()
    driver.quit()


@pytest.mark.usefixtures("log_on_failure")
@pytest.mark.parametrize("username, password", get_data2())
def test_dologin(username, password, get_browser):
    driver = get_browser
    driver.find_element(By.ID, "email").send_keys(username)
    driver.find_element(By.ID, "pass").send_keys(password)
    #assert 1 == 2
    # allure.attach(driver.get_screenshot_as_png(), name='dologin',attachment_type=AttachmentType.PNG)

# run parallel mode install -->  pip install pytest-xdist or add addopts = -n3 <--in pytest.ini file
# run the test with -->  pytest test_webdriver_integration.py -n 3 <-- where 3 is nuber of parallel tests
# run --> pytest test_webdriver_integration.py -n 3 --alluredir="./allurereports" <--to generate allure json
# run --> allure serve ./allurereports
